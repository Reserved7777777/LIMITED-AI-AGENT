#!/usr/bin/env python3
"""
Fetch all A-share stocks with industry classification and prices.
Uses East Money API's board membership to determine industry.
Output: CSV + Excel (.xlsx)
"""

import requests
import csv
import json
import time
import sys


def get_all_industry_boards():
    """Get all SW industry boards (m:90+t:2) with pagination"""
    url = "http://push2.eastmoney.com/api/qt/clist/get"
    headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://quote.eastmoney.com/"}
    
    all_boards = []
    page = 1
    while True:
        params = {
            "pn": page, "pz": 100,
            "po": 1, "np": 1,
            "ut": "bd1d9ddb04089700cf9c27f6f7426281",
            "fltt": 2, "invt": 2, "fid": "f12",
            "fs": "m:90+t:2",
            "fields": "f12,f14"
        }
        try:
            resp = requests.get(url, params=params, headers=headers, timeout=15)
            data = resp.json()
            if data.get("data") and data["data"].get("diff"):
                batch = data["data"]["diff"]
                if not batch:
                    break
                all_boards.extend(batch)
                page += 1
            else:
                break
        except Exception as e:
            print(f"  Board page error: {e}")
            break
    
    print(f"Industry boards: {len(all_boards)}")
    return all_boards


def get_board_stocks(board_code):
    """Get all stocks in a specific industry board"""
    url = "http://push2.eastmoney.com/api/qt/clist/get"
    headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://quote.eastmoney.com/"}
    
    all_stocks = []
    page = 1
    total = None
    
    while True:
        params = {
            "pn": page, "pz": 500,
            "po": 1, "np": 1,
            "ut": "bd1d9ddb04089700cf9c27f6f7426281",
            "fltt": 2, "invt": 2,
            "fs": f"b:{board_code}+f:!50",
            "fields": "f12,f14"
        }
        try:
            resp = requests.get(url, params=params, headers=headers, timeout=15)
            data = resp.json()
            if data.get("data") and data["data"].get("diff"):
                batch = data["data"]["diff"]
                if total is None:
                    total = data["data"].get("total", 0)
                all_stocks.extend(batch)
                if len(all_stocks) >= total:
                    break
                page += 1
            else:
                break
        except Exception as e:
            print(f"    Board {board_code} page error: {e}")
            break
    
    return all_stocks


def get_all_stocks_with_prices():
    """Get all A-share stocks with basic info (code, name, price)"""
    url = "http://push2.eastmoney.com/api/qt/clist/get"
    headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://quote.eastmoney.com/"}
    
    all_stocks = {}
    page = 1
    total = None
    
    while True:
        params = {
            "pn": page, "pz": 2000,
            "po": 0, "np": 1,
            "ut": "bd1d9ddb04089700cf9c27f6f7426281",
            "fltt": 2, "invt": 2, "fid": "f12",
            "fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
            "fields": "f2,f3,f12,f14,f20,f21"
        }
        try:
            resp = requests.get(url, params=params, headers=headers, timeout=30)
            data = resp.json()
            if data.get("data") and data["data"].get("diff"):
                batch = data["data"]["diff"]
                if total is None:
                    total = data["data"].get("total", 0)
                for s in batch:
                    code = s.get("f12", "")
                    all_stocks[code] = s
                print(f"  Stocks page: +{len(batch)} (total: {len(all_stocks)})")
                if len(all_stocks) >= total:
                    break
                page += 1
            else:
                break
        except Exception as e:
            print(f"  Page error: {e}")
            break
    
    return all_stocks


def format_market_cap(val):
    if val is None or val in ("-", ""):
        return "-"
    try:
        v = float(val)
        if v >= 1e8:
            return f"{v/1e8:.2f}亿"
        else:
            return f"{v:.2f}"
    except:
        return str(val)


def main():
    print("=" * 60)
    print("Step 1: Fetching all industry boards...")
    print("=" * 60)
    boards = get_all_industry_boards()
    if not boards:
        print("ERROR: No industry boards!")
        return
    
    board_list = [(b["f12"], b["f14"]) for b in boards]
    print(f"Total boards: {len(board_list)}")
    
    print("\n" + "=" * 60)
    print("Step 2: Fetching member stocks for each board...")
    print("=" * 60)
    
    # Build stock -> industry mapping
    # stock_code -> list of (board_code, board_name)
    stock_industries = {}
    
    for idx, (bk_code, bk_name) in enumerate(board_list):
        stocks = get_board_stocks(bk_code)
        if stocks:
            for s in stocks:
                code = s.get("f12", "")
                if code not in stock_industries:
                    stock_industries[code] = []
                # Use the SW level-3 industry name (most specific)
                stock_industries[code].append(bk_name)
        
        if (idx + 1) % 10 == 0:
            print(f"  Processed {idx+1}/{len(board_list)} boards, mapped {len(stock_industries)} stocks")
        
        time.sleep(0.15)  # Rate limiting
    
    print(f"\nMapped {len(stock_industries)} unique stocks to industries")
    
    print("\n" + "=" * 60)
    print("Step 3: Getting price data for all stocks...")
    print("=" * 60)
    
    all_stocks = get_all_stocks_with_prices()
    if not all_stocks:
        print("ERROR: No stock data!")
        return
    
    print(f"Total stocks with prices: {len(all_stocks)}")
    
    # Merge data
    print("\n" + "=" * 60)
    print("Step 4: Creating output files...")
    print("=" * 60)
    
    # Write CSV
    csv_path = "/root/.openclaw/workspace/A股全量数据.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["股票代码", "股票名称", "所属行业(申万)", "最新价", "涨跌幅(%)", "总市值", "流通市值"])
        
        for code in sorted(all_stocks.keys()):
            s = all_stocks[code]
            name = s.get("f14", "")
            
            # Get industry - use the most specific SW level-3 industry
            industries = stock_industries.get(code, [])
            # Choose the industry with the most specific name (longest, or the first one)
            industry = industries[0] if industries else ""
            
            price = s.get("f2", "")
            change_pct = s.get("f3", "")
            mcap = format_market_cap(s.get("f20"))
            float_mcap = format_market_cap(s.get("f21"))
            
            writer.writerow([code, name, industry, price, change_pct, mcap, float_mcap])
    
    print(f"CSV: {csv_path} ({len(all_stocks)} rows)")
    
    # Write XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "A股全量数据"
        
        headers = ["股票代码", "股票名称", "所属行业(申万)", "最新价", "涨跌幅(%)", "总市值", "流通市值"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        row_idx = 2
        for code in sorted(all_stocks.keys()):
            s = all_stocks[code]
            name = s.get("f14", "")
            industries = stock_industries.get(code, [])
            industry = industries[0] if industries else ""
            price = s.get("f2", "")
            change_pct = s.get("f3", "")
            mcap = format_market_cap(s.get("f20"))
            float_mcap = format_market_cap(s.get("f21"))
            
            row_data = [code, name, industry, price, change_pct, mcap, float_mcap]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=10, name="微软雅黑")
                if col == 1:
                    cell.alignment = Alignment(horizontal='center')
                elif col == 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 5:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            row_idx += 1
            if row_idx % 1000 == 0:
                print(f"  Writing row {row_idx}...")
        
        widths = [14, 22, 28, 12, 12, 16, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{len(all_stocks)+1}"
        
        xlsx_path = "/root/.openclaw/workspace/A股全量数据.xlsx"
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
        
    except ImportError:
        print("openpyxl not available - CSV only")
    except Exception as e:
        print(f"Excel error: {e}")
    
    # Stats
    mapped_count = sum(1 for code in all_stocks if stock_industries.get(code))
    print(f"\n{'='*60}")
    print(f"Summary: {len(all_stocks)} stocks, {mapped_count} with industry info")
    print(f"Files updated: {csv_path}, XLSX")
    
    # Show some stats about industry distribution
    industry_counts = {}
    for code in all_stocks:
        inds = stock_industries.get(code, [])
        ind = inds[0] if inds else "未知"
        industry_counts[ind] = industry_counts.get(ind, 0) + 1
    
    print(f"\nTop 20 Industries:")
    for ind, cnt in sorted(industry_counts.items(), key=lambda x: -x[1])[:20]:
        print(f"  {ind}: {cnt}")
    
    print(f"\nUnknown: {industry_counts.get('未知', 0)}")


if __name__ == "__main__":
    main()
