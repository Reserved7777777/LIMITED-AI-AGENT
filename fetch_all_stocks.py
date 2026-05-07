#!/usr/bin/env python3
"""
Fetch all A-share stocks: code, name, industry (申万行业), current price
Output: CSV + Excel (.xlsx)
"""

import requests
import csv
import json
import time
import re


def get_stock_market_prefix(code):
    """Determine market prefix for East Money"""
    code_str = str(code).zfill(6)
    if code_str.startswith('6') or code_str.startswith('9'):
        return '1'  # Shanghai
    else:
        return '0'  # Shenzhen


def fetch_all_stocks(page_size=2000):
    """Fetch all A-share stocks with basic info"""
    all_stocks = []
    page = 1
    total = None

    url = "http://push2.eastmoney.com/api/qt/clist/get"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://quote.eastmoney.com/"
    }

    while True:
        params = {
            "pn": page,
            "pz": page_size,
            "po": 0,  # sort ascending by fid
            "np": 1,
            "ut": "bd1d9ddb04089700cf9c27f6f7426281",
            "fltt": 2,
            "invt": 2,
            "fid": "f12",
            "fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
            # f2=price, f3=change%, f12=code, f14=name, f20=total_cap, f21=float_cap
            # f30=industry code (SW申万), f37=industry name (申万三级)
            "fields": "f2,f3,f12,f14,f20,f21,f30,f37"
        }

        try:
            resp = requests.get(url, params=params, headers=headers, timeout=30)
            data = resp.json()
            if not data.get("data") or not data["data"].get("diff"):
                break

            batch = data["data"]["diff"]
            if total is None:
                total = data["data"].get("total", 0)
                print(f"Total stocks: {total}")

            all_stocks.extend(batch)
            print(f"  Page {page}: +{len(batch)} (total: {len(all_stocks)})")

            if len(all_stocks) >= total:
                break

            page += 1
            time.sleep(0.2)

        except Exception as e:
            print(f"  Page {page} error: {e}")
            time.sleep(1)
            page += 1

    return all_stocks


def fetch_sw_industry_map():
    """
    Fetch SW (申万) industry classification from East Money
    The f37 field in stock list gives SW level-3 industry name.
    But sometimes it's empty. Let's try to get industry mapping.
    """
    url = "http://push2.eastmoney.com/api/qt/clist/get"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://quote.eastmoney.com/"
    }

    # First get industry sectors (SW level-3) - m:90+t:2
    params = {
        "pn": 1, "pz": 600,
        "po": 1, "np": 1,
        "ut": "bd1d9ddb04089700cf9c27f6f7426281",
        "fltt": 2, "invt": 2, "fid": "f12",
        "fs": "m:90+t:2",
        "fields": "f12,f14"
    }

    try:
        resp = requests.get(url, params=params, headers=headers, timeout=15)
        data = resp.json()
        if data.get("data") and data["data"].get("diff"):
            boards = {}
            for b in data["data"]["diff"]:
                bk_code = b.get("f12", "")
                bk_name = b.get("f14", "")
                boards[bk_code] = bk_name
            print(f"Found {len(boards)} industry boards")
            return boards
    except Exception as e:
        print(f"Error fetching industry boards: {e}")

    return {}


def get_stock_industry_via_board(stock_code, board_data):
    """
    Determine industry for a stock by checking which board it belongs to.
    This requires many API calls... not practical for 5300 stocks.
    Alternative: use f37 or f30 fields from stock data.
    """
    pass


def format_value(val):
    """Format a stock data value"""
    if val is None or val == "-" or val == "":
        return "-"
    return val


def format_market_cap(val):
    if val is None or val == "-" or val == "":
        return "-"
    try:
        v = float(val)
        if v >= 1e8:
            return f"{v/1e8:.2f}亿"
        else:
            return f"{v:.2f}"
    except:
        return str(val)


def main():
    print("=" * 60)
    print("Fetching all A-share stocks from East Money...")
    print("=" * 60)

    stocks = fetch_all_stocks()
    if not stocks:
        print("ERROR: No stocks fetched!")
        return

    print(f"\nTotal: {len(stocks)} stocks")

    # Try to get industry name from f37 (SW industry level-3)
    # f37 might not be what we think, let's check a few samples
    print("\n--- Sample stocks with f37 (industry?) ---")
    for s in stocks[:20]:
        code = s.get("f12", "")
        name = s.get("f14", "")
        f30 = s.get("f30", "")
        f37 = s.get("f37", "")
        print(f"  {code} | {name} | f30={f30} | f37={f37}")    

    # Check how many have f37 populated
    with_f37 = sum(1 for s in stocks if s.get("f37") and str(s.get("f37", "")).strip() not in ("", "-"))
    print(f"\nStocks with industry (f37): {with_f37}/{len(stocks)}")

    # Strategy: Use f30 as industry code and try to create a mapping
    # Check unique f30 values
    all_f30 = set()
    for s in stocks:
        f30 = s.get("f30")
        if f30 is not None:
            all_f30.add(str(f30))
    print(f"Unique f30 industry codes: {len(all_f30)}")

    # Output CSV
    csv_path = "/root/.openclaw/workspace/A股全量数据.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["股票代码", "股票名称", "所属行业", "最新价", "涨跌幅(%)", "总市值", "流通市值"])

        for s in stocks:
            code = format_value(s.get("f12"))
            name = format_value(s.get("f14"))
            industry = format_value(s.get("f37", ""))
            price = format_value(s.get("f2"))
            change_pct = format_value(s.get("f3"))
            mcap = format_market_cap(s.get("f20"))
            float_mcap = format_market_cap(s.get("f21"))

            writer.writerow([code, name, industry, price, change_pct, mcap, float_mcap])

    print(f"\nCSV saved: {csv_path}")

    # Try to create XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "A股全量数据"

        headers = ["股票代码", "股票名称", "所属行业", "最新价", "涨跌幅(%)", "总市值", "流通市值"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx, s in enumerate(stocks, 2):
            code = format_value(s.get("f12"))
            name = format_value(s.get("f14"))
            industry = format_value(s.get("f37", ""))
            price = format_value(s.get("f2"))
            change_pct = format_value(s.get("f3"))
            mcap = format_market_cap(s.get("f20"))
            float_mcap = format_market_cap(s.get("f21"))

            row_data = [code, name, industry, price, change_pct, mcap, float_mcap]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=10, name="微软雅黑")
                if col == 1:
                    cell.alignment = Alignment(horizontal='center')
                elif col == 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 5:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='right')

            if row_idx % 1000 == 0:
                print(f"  Writing row {row_idx}...")

        widths = [14, 20, 24, 12, 12, 16, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{len(stocks)+1}"

        xlsx_path = "/root/.openclaw/workspace/A股全量数据.xlsx"
        wb.save(xlsx_path)
        print(f"Excel saved: {xlsx_path}")

    except ImportError:
        print("openpyxl not available - CSV only. Install with: pip install openpyxl")
    except Exception as e:
        print(f"Excel creation error: {e}")

    # Final stats
    print(f"\n{'='*60}")
    print(f"Complete! Total stocks: {len(stocks)}")
    print(f"Files:")
    print(f"  CSV: {csv_path}")
    print(f"  XLSX: /root/.openclaw/workspace/A股全量数据.xlsx (if created)")

    # Industry distribution (top 20)
    industries = {}
    for s in stocks:
        ind = str(s.get("f37", "") or "").strip()
        if not ind or ind == "-":
            ind = "未知"
        industries[ind] = industries.get(ind, 0) + 1

    print(f"\n=== Industry Distribution (top 30) ===")
    for ind, count in sorted(industries.items(), key=lambda x: -x[1])[:30]:
        print(f"  {ind}: {count}")


if __name__ == "__main__":
    main()
