#!/usr/bin/env python3
"""
Fetch all A-share stocks (code, name, industry, current price) from East Money API.
Output: CSV + Excel (.xlsx)
"""

import requests
import csv
import time
import json

def fetch_a_shares(page_size=5500):
    """Fetch all A-share stocks (Shanghai + Shenzhen)"""
    url = "http://push2.eastmoney.com/api/qt/clist/get"
    params = {
        "pn": 1,
        "pz": page_size,
        "po": 1,
        "np": 1,
        "ut": "bd1d9ddb04089700cf9c27f6f7426281",
        "fltt": 2,
        "invt": 2,
        "fid": "f12",
        "fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
        # f2=现价, f3=涨跌幅, f12=代码, f14=名称, f20=总市值, f21=流通市值
        # f30=行业代码, f31=行业名称, f33=板块名称
        "fields": "f2,f3,f12,f14,f20,f21,f30,f31,f33"
    }

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://quote.eastmoney.com/"
    }

    try:
        resp = requests.get(url, params=params, headers=headers, timeout=30)
        data = resp.json()
        if data and data.get("data") and data["data"].get("diff"):
            stocks = data["data"]["diff"]
            total = data["data"].get("total", 0)
            print(f"Total: {total}, Fetched: {len(stocks)}")
            return stocks
        else:
            print(f"API returned unexpected structure: {json.dumps(data, ensure_ascii=False)[:500]}")
            return []
    except Exception as e:
        print(f"Error fetching: {e}")
        return []


def format_market_cap(val):
    """Format market cap to readable string"""
    if val is None or val == "-":
        return "-"
    try:
        v = float(val)
        if v >= 1e12:
            return f"{v/1e12:.2f}万亿"
        elif v >= 1e8:
            return f"{v/1e8:.2f}亿"
        elif v >= 1e4:
            return f"{v/1e4:.2f}万"
        else:
            return str(v)
    except:
        return str(val)


def main():
    # Step 1: Try fetching all at once
    print("Fetching all A-share stocks from East Money...")
    stocks = fetch_a_shares()

    if not stocks:
        print("No data returned. Trying fallback with paging...")
        # Fallback: page through
        stocks = []
        for page in range(1, 6):
            url = "http://push2.eastmoney.com/api/qt/clist/get"
            params = {
                "pn": page, "pz": 1000, "po": 1, "np": 1,
                "ut": "bd1d9ddb04089700cf9c27f6f7426281",
                "fltt": 2, "invt": 2, "fid": "f12",
                "fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
                "fields": "f2,f3,f12,f14,f20,f21,f30,f31,f33"
            }
            try:
                resp = requests.get(url, params=params, headers={
                    "User-Agent": "Mozilla/5.0",
                    "Referer": "https://quote.eastmoney.com/"
                }, timeout=30)
                data = resp.json()
                if data.get("data", {}).get("diff"):
                    batch = data["data"]["diff"]
                    stocks.extend(batch)
                    print(f"  Page {page}: +{len(batch)} stocks")
                time.sleep(0.3)
            except Exception as e:
                print(f"  Page {page} error: {e}")

    if not stocks:
        print("Still no data. Trying alternative API endpoint...")
        # Alternative: Sina Finance API for stock listing
        # This is a different approach - fetch all stock codes first
        print("Aborting - no data source available.")
        return

    # Step 2: Write CSV
    csv_path = "/root/.openclaw/workspace/A股全量数据.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["股票代码", "股票名称", "所属行业", "最新价", "涨跌幅(%)", "总市值", "流通市值"])

        for s in stocks:
            code = s.get("f12", "")
            name = s.get("f14", "")
            # Try multiple industry fields
            industry = s.get("f31") or s.get("f33") or s.get("f30") or ""
            price = s.get("f2", "")
            change_pct = s.get("f3", "")
            mcap = format_market_cap(s.get("f20"))
            float_mcap = format_market_cap(s.get("f21"))

            writer.writerow([code, name, industry, price, change_pct, mcap, float_mcap])

    total = len(stocks)
    print(f"\nSaved {total} stocks to {csv_path}")

    # Step 3: Try to create xlsx if openpyxl is available
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "A股全量数据"

        # Headers
        headers = ["股票代码", "股票名称", "所属行业", "最新价", "涨跌幅(%)", "总市值", "流通市值"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data rows
        for row_idx, s in enumerate(stocks, 2):
            code = s.get("f12", "")
            name = s.get("f14", "")
            industry = s.get("f31") or s.get("f33") or s.get("f30") or ""
            price = s.get("f2", "")
            change_pct = s.get("f3", "")
            mcap = format_market_cap(s.get("f20"))
            float_mcap = format_market_cap(s.get("f21"))

            row_data = [code, name, industry, price, change_pct, mcap, float_mcap]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col == 1:
                    cell.alignment = Alignment(horizontal='center')
                elif col == 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 5:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='right')

        # Column widths
        widths = [14, 18, 20, 12, 12, 16, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Freeze top row
        ws.freeze_panes = "A2"

        xlsx_path = "/root/.openclaw/workspace/A股全量数据.xlsx"
        wb.save(xlsx_path)
        print(f"Also saved Excel: {xlsx_path}")

    except ImportError:
        print("openpyxl not available, CSV only.")

    print(f"\nDone! Total: {total} stocks")

    # Show sample
    print("\n=== Sample (first 5 rows) ===")
    for s in stocks[:5]:
        code = s.get("f12", "")
        name = s.get("f14", "")
        industry = s.get("f31") or s.get("f33") or s.get("f30") or ""
        price = s.get("f2", "")
        print(f"  {code} | {name} | {industry} | {price}")

    # Show industry stats
    industries = {}
    for s in stocks:
        ind = s.get("f31") or s.get("f33") or s.get("f30") or "未知"
        industries[ind] = industries.get(ind, 0) + 1
    print(f"\n=== Industry Count ({len(industries)} industries) ===")
    for ind, count in sorted(industries.items(), key=lambda x: -x[1])[:20]:
        print(f"  {ind}: {count}")


if __name__ == "__main__":
    main()
